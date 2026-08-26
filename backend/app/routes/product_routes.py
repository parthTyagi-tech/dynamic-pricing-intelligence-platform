import csv
import io
from flask import Blueprint, request, send_file
from flask_jwt_extended import (
    jwt_required,
    get_jwt_identity
)
from app.extensions import db
from app.models.user import User
from app.models.product import Product
from app.utils.decorators import admin_required


product_bp = Blueprint(
    "products",
    __name__
)


# =====================================
# CREATE PRODUCT
# =====================================

@product_bp.route(
    "",
    methods=["POST"]
)
@admin_required()
def create_product():

    current_user_id = get_jwt_identity()

    current_user = User.query.get(current_user_id)

    data = request.get_json()

    required_fields = [
        
        "name",
        
        "current_price",
        
    ]

    for field in required_fields:

        if field not in data:

            return {
                "success": False,
                "message": f"{field} is required"
            }, 400

    existing_product = Product.query.filter_by(
        sku=data.get("sku", "AUTO-SKU"),
    ).first()

    if existing_product:

        return {
            "success": False,
            "message": "SKU already exists"
        }, 400

    product = Product(
        sku=data["sku"],
        name=data["name"],
        category=data.get("category", "General"),
        brand=data.get("brand"),
        barcode=data.get("barcode"),
        description=data.get("description"),
        current_price=data["current_price"],
        cost_price=data["cost_price"],
        inventory_quantity=data.get(
            "inventory_quantity",
            0
        ),
        organization_id=current_user.organization_id
    )

    db.session.add(product)

    db.session.commit()

    return {
        "success": True,
        "message": "Product created successfully",
        "product": product.to_dict()
    }, 201


# =====================================
# GET ALL PRODUCTS
# =====================================

@product_bp.route(
    "",
    methods=["GET"]
)
@jwt_required()
def get_products():

    current_user_id = get_jwt_identity()

    current_user = User.query.get(current_user_id)

    products = Product.query.filter_by(
        organization_id=current_user.organization_id
    ).all()

    return {
        "success": True,
        "count": len(products),
        "products": [
            product.to_dict()
            for product in products
        ]
    }, 200


# =====================================
# GET SINGLE PRODUCT
# =====================================

@product_bp.route(
    "/<product_id>",
    methods=["GET"]
)
@jwt_required()
def get_product(product_id):

    current_user_id = get_jwt_identity()

    current_user = User.query.get(current_user_id)

    product = Product.query.filter_by(
        id=product_id,
        organization_id=current_user.organization_id
    ).first()

    if not product:

        return {
            "success": False,
            "message": "Product not found"
        }, 404

    return {
        "success": True,
        "product": product.to_dict()
    }, 200


# =====================================
# UPDATE PRODUCT
# =====================================

@product_bp.route(
    "/<product_id>",
    methods=["PUT"]
)
@admin_required()
def update_product(product_id):

    current_user_id = get_jwt_identity()

    current_user = User.query.get(current_user_id)

    product = Product.query.filter_by(
        id=product_id,
        organization_id=current_user.organization_id
    ).first()

    if not product:

        return {
            "success": False,
            "message": "Product not found"
        }, 404

    data = request.get_json()

    product.name = data.get(
        "name",
        product.name
    )

    product.category = data.get(
        "category",
        product.category
    )

    product.description = data.get(
        "description",
        product.description
    )

    product.current_price = data.get(
        "current_price",
        product.current_price
    )

    product.cost_price = data.get(
        "cost_price",
        product.cost_price
    )

    product.inventory_quantity = data.get(
        "inventory_quantity",
        product.inventory_quantity
    )

    db.session.commit()

    return {
        "success": True,
        "message": "Product updated successfully",
        "product": product.to_dict()
    }, 200


# =====================================
# DELETE PRODUCT
# =====================================

@product_bp.route(
    "/<product_id>",
    methods=["DELETE"]
)
@admin_required()
def delete_product(product_id):

    current_user_id = get_jwt_identity()

    current_user = User.query.get(current_user_id)

    product = Product.query.filter_by(
        id=product_id,
        organization_id=current_user.organization_id
    ).first()

    if not product:

        return {
            "success": False,
            "message": "Product not found"
        }, 404

    db.session.delete(product)

    db.session.commit()

    return {
        "success": True,
        "message": "Product deleted successfully"
    }, 200

# =====================================
# IMPORT PRODUCTS VIA CSV
# =====================================

@product_bp.route(
    "/import-csv",
    methods=["POST"]
)
@jwt_required()
def import_csv():
    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)

    if not current_user:
        return {
            "success": False,
            "message": "User not found"
        }, 404

    # Check if a file was uploaded
    if "file" not in request.files:
        return {
            "success": False,
            "message": "No file uploaded"
        }, 400

    file = request.files["file"]
    if file.filename == "":
        return {
            "success": False,
            "message": "No file selected"
        }, 400

    if not file.filename.lower().endswith((".csv", ".xlsx")):
        return {
            "success": False,
            "message": "File format must be CSV or XLSX"
        }, 400

    try:
        import io
        import csv
        import uuid

        if file.filename.lower().endswith(".xlsx"):
            from openpyxl import load_workbook
            workbook = load_workbook(io.BytesIO(file.stream.read()), read_only=True, data_only=True)
            worksheet = workbook.active
            rows = worksheet.iter_rows(values_only=True)
            header_row = next(rows, ())
            headers = [str(value or "").strip().lower() for value in header_row]
            reader = (dict(zip(headers, values)) for values in rows)
        else:
            # Read the file stream with utf-8-sig to strip byte order mark (BOM) if present
            stream = io.StringIO(file.stream.read().decode("utf-8-sig"), newline=None)
            csv_reader = csv.DictReader(stream)
            headers = [str(header or "").strip().lower() for header in (csv_reader.fieldnames or [])]
            reader = csv_reader

        # Validate headers
        required_headers = ["name", "current_price"]
        
        for req in required_headers:
            if req not in headers:
                return {
                    "success": False,
                    "message": f"Missing required column: '{req}'. File must contain at least 'name' and 'current_price' columns."
                }, 400

        products_imported = 0
        db_mappings = []
        seen_skus_in_batch = {}

        # Read row-by-row
        for row in reader:
            # Map headers case-insensitively
            row_clean = {str(k).strip().lower(): str(v).strip() if v is not None else "" for k, v in row.items() if k}
            
            name = row_clean.get("name")
            current_price_str = row_clean.get("current_price")
            
            if not name or not current_price_str:
                continue

            try:
                current_price = float(current_price_str)
                cost_price = float(row_clean.get("cost_price", 0.0) or 0.0)
            except ValueError:
                continue

            sku = row_clean.get("sku") or f"SKU-{str(uuid.uuid4())[:8].upper()}"
            category = row_clean.get("category", "General")
            description = row_clean.get("description", "")
            brand = row_clean.get("brand", "")
            barcode = row_clean.get("barcode", "")
            inventory_qty = int(row_clean.get("inventory_quantity", 0) or 0)

            category_hint = row_clean.get("category_hint") or category
            normalized_query = row_clean.get("normalized_query") or name
            
            # Parse attributes JSON if present
            attrs_val = {}
            if "attributes" in row_clean and row_clean["attributes"]:
                try:
                    import json
                    attrs_val = json.loads(row_clean["attributes"])
                    if not isinstance(attrs_val, dict):
                        attrs_val = {}
                except Exception:
                    attrs_val = {}

            # Check if SKU already exists in this organization
            existing = Product.query.filter_by(sku=sku, organization_id=current_user.organization_id).first()
            if existing:
                existing.name = name
                existing.current_price = current_price
                existing.cost_price = cost_price
                existing.category = category
                existing.brand = brand
                existing.barcode = barcode
                existing.description = description
                existing.inventory_quantity = inventory_qty
                existing.category_hint = category_hint
                existing.normalized_query = normalized_query
                existing.attributes = attrs_val
            elif sku in seen_skus_in_batch:
                # Update the already mapped object in memory instead of inserting again
                mapping = seen_skus_in_batch[sku]
                mapping["name"] = name
                mapping["current_price"] = current_price
                mapping["cost_price"] = cost_price
                mapping["category"] = category
                mapping["brand"] = brand
                mapping["barcode"] = barcode
                mapping["description"] = description
                mapping["inventory_quantity"] = inventory_qty
                mapping["category_hint"] = category_hint
                mapping["normalized_query"] = normalized_query
                mapping["attributes"] = attrs_val
            else:
                new_mapping = {
                    "sku": sku,
                    "name": name,
                    "category": category,
                    "brand": brand,
                    "barcode": barcode,
                    "description": description,
                    "current_price": current_price,
                    "cost_price": cost_price,
                    "inventory_quantity": inventory_qty,
                    "organization_id": current_user.organization_id,
                    "category_hint": category_hint,
                    "normalized_query": normalized_query,
                    "attributes": attrs_val
                }
                db_mappings.append(new_mapping)
                seen_skus_in_batch[sku] = new_mapping
            
            products_imported += 1

        if db_mappings:
            db.session.bulk_insert_mappings(Product, db_mappings)
        
        db.session.commit()

        return {
            "success": True,
            "message": f"Successfully imported {products_imported} products",
            "imported_count": products_imported
        }, 200

    except Exception as e:
        db.session.rollback()
        print(f"[import_csv] Error: {e}")
        return {
            "success": False,
            "message": f"Failed to process CSV file: {str(e)}"
        }, 500


# =====================================
# EXPORT STORED CATALOG WITH CURRENT PRICES
# =====================================

@product_bp.route("/export-csv", methods=["GET"])
@jwt_required()
def export_catalog():
    current_user = User.query.get(get_jwt_identity())
    if not current_user:
        return {"success": False, "message": "User not found"}, 404

    file_format = (request.args.get("format") or request.args.get("file_type") or "csv").lower()
    if file_format not in {"csv", "xlsx"}:
        return {"success": False, "message": "format must be csv or xlsx"}, 400

    products = Product.query.filter_by(organization_id=current_user.organization_id).order_by(Product.sku.asc()).all()
    headers = ["sku", "name", "brand", "barcode", "category", "description", "current_price", "cost_price", "inventory_quantity", "min_margin_percentage", "recommendation_status"]
    rows = [[
        product.sku,
        product.name,
        product.brand or "",
        product.barcode or "",
        product.category,
        product.description or "",
        product.current_price,
        product.cost_price,
        product.inventory_quantity,
        product.min_margin_percentage,
        product.recommendation_status,
    ] for product in products]

    if file_format == "xlsx":
        from openpyxl import Workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Catalog"
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="klypup-catalog-updated.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    return send_file(io.BytesIO(output.getvalue().encode("utf-8-sig")), as_attachment=True, download_name="klypup-catalog-updated.csv", mimetype="text/csv")
